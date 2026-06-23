import logging
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)
REPORTS_DIR = Path(__file__).resolve().parent.parent / "reports"

REPORT_COLUMNS = [
    ("№", None),
    ("Название Ozon", "ozon_title"),
    ("Название Kaspi", "kaspi_title"),
    ("Бренд", "brand"),
    ("Цена Ozon", "ozon_price"),
    ("Цена Kaspi", "kaspi_price"),
    ("Доставка", "delivery"),
    ("Себестоимость", "total_cost"),
    ("Чистая выручка", "net_revenue"),
    ("Прибыль", "profit"),
    ("ROI %", "roi"),
    ("Match Score", "match_score"),
    ("Ссылка Ozon", "ozon_url"),
    ("Ссылка Kaspi", "kaspi_url"),
]

MONEY_COLUMNS = {5, 6, 7, 8, 9, 10}
LINK_COLUMNS = {13, 14}

INTERNET_REPORT_COLUMNS = [
    ("№", None),
    ("Название Ozon", "ozon_title"),
    ("Название в магазине", "internet_title"),
    ("Бренд", "brand"),
    ("Модель", "model"),
    ("Источник", "source"),
    ("Источников с ценой", "sources_count"),
    ("Цена Ozon", "ozon_price"),
    ("Цена в интернете", "internet_price"),
    ("Комиссия %", "commission_rate"),
    ("Комиссия", "commission"),
    ("Чистая выручка", "net_revenue"),
    ("Доставка", "delivery"),
    ("Себестоимость", "total_cost"),
    ("Разница цен", "price_difference"),
    ("Прибыль", "profit"),
    ("ROI %", "roi"),
    ("Match Score", "match_score"),
    ("Наличие", "availability"),
    ("Ссылка Ozon", "ozon_url"),
    ("Ссылка магазина", "internet_url"),
]

INTERNET_MONEY_COLUMNS = {8, 9, 11, 12, 13, 14, 15, 16}
INTERNET_LINK_COLUMNS = {20, 21}


def save_arbitrage_report(items: list[dict]) -> str:
    """Сохраняет результаты арбитража в Excel и возвращает путь."""
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    filename = datetime.now().strftime("arbitrage_%Y_%m_%d_%H_%M.xlsx")
    report_path = REPORTS_DIR / filename

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Arbitrage"
    sheet.freeze_panes = "A2"

    header_fill = PatternFill(
        fill_type="solid",
        start_color="1F4E78",
        end_color="1F4E78",
    )
    header_font = Font(color="FFFFFF", bold=True)
    for column, (title, _) in enumerate(REPORT_COLUMNS, 1):
        cell = sheet.cell(row=1, column=column, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for row_index, item in enumerate(items, 2):
        sheet.cell(row=row_index, column=1, value=row_index - 1)
        for column, (_, key) in enumerate(REPORT_COLUMNS[1:], 2):
            value: Any = item.get(key) if key else None
            cell = sheet.cell(row=row_index, column=column, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if column in MONEY_COLUMNS and value is not None:
                cell.number_format = '#,##0.00 "₸"'
            elif column in {11, 12} and value is not None:
                cell.number_format = "0.00"
            elif column in LINK_COLUMNS and value:
                cell.hyperlink = str(value)
                cell.style = "Hyperlink"

    widths = {
        1: 6,
        2: 45,
        3: 45,
        4: 20,
        5: 15,
        6: 15,
        7: 14,
        8: 16,
        9: 18,
        10: 15,
        11: 12,
        12: 14,
        13: 50,
        14: 50,
    }
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = width

    last_row = max(sheet.max_row, 1)
    sheet.auto_filter.ref = f"A1:N{last_row}"
    sheet.row_dimensions[1].height = 24

    workbook.save(report_path)
    logger.info("Excel-отчет арбитража сохранен: %s", report_path)
    return str(report_path)


def save_internet_comparison_report(items: list[dict]) -> str:
    """Сохраняет сравнение Ozon с интернет-ценами в Excel."""
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    filename = datetime.now().strftime(
        "internet_comparison_%Y_%m_%d_%H_%M.xlsx"
    )
    report_path = REPORTS_DIR / filename

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ozon vs Internet"
    sheet.freeze_panes = "A2"

    header_fill = PatternFill(
        fill_type="solid",
        start_color="E85D04",
        end_color="E85D04",
    )
    header_font = Font(color="FFFFFF", bold=True)
    for column, (title, _) in enumerate(INTERNET_REPORT_COLUMNS, 1):
        cell = sheet.cell(row=1, column=column, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for row_index, item in enumerate(items, 2):
        sheet.cell(row=row_index, column=1, value=row_index - 1)
        for column, (_, key) in enumerate(
            INTERNET_REPORT_COLUMNS[1:],
            2,
        ):
            value: Any = item.get(key) if key else None
            cell = sheet.cell(row=row_index, column=column, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if column in INTERNET_MONEY_COLUMNS and value is not None:
                cell.number_format = '#,##0.00 "₸"'
            elif column in {10, 17, 18} and value is not None:
                cell.number_format = "0.00"
            elif column in INTERNET_LINK_COLUMNS and value:
                cell.hyperlink = str(value)
                cell.style = "Hyperlink"

    widths = {
        1: 6,
        2: 45,
        3: 45,
        4: 18,
        5: 18,
        6: 22,
        7: 18,
        8: 15,
        9: 17,
        10: 14,
        11: 15,
        12: 18,
        13: 16,
        14: 16,
        15: 16,
        16: 15,
        17: 12,
        18: 14,
        19: 20,
        20: 50,
        21: 50,
    }
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = width

    last_row = max(sheet.max_row, 1)
    sheet.auto_filter.ref = f"A1:U{last_row}"
    sheet.row_dimensions[1].height = 24

    workbook.save(report_path)
    logger.info(
        "Excel-отчет интернет-сравнения сохранен: %s",
        report_path,
    )
    return str(report_path)
